from flask import Flask, request, render_template, send_file,jsonify,flash,redirect, url_for

from openpyxl import load_workbook
import pandas as pd
from datetime import date, timedelta
import datetime

from gevent.pywsgi import WSGIServer

from os import system, name 

import smtplib, ssl
import os

from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart


import random, threading, webbrowser




un=''   #username
e=''    #email id
d=''    #department
yrs=''  #year and section
ro=''   #row
password=''   #password
erow=[]  #Entry row
found=False  
bookl=[]   #books list
rl=[]  #renewal list
rrow=[]  #retur row indexs
rbd=[]  #renewal book due dates
foundr=False #data found in return
rbookl=[]  #booklist

filename=r'files\credentials.xlsx'
filename1 = r'files\Entry.xlsx'


app = Flask(__name__, template_folder='templates')
app.secret_key = 'super secret'


@app.route('/')
def login_page():
    library_line() 
    return render_template('login.html')

@app.route('/assets/<path:path>')
def serve_static(path):
    return send_file("assets/"+path)


@app.route('/register')
def register_page():
    return render_template('register.html')

@app.route('/forgot')
def forgot_page():
    return render_template('forgot.html')

@app.route('/sforgot' ,methods=['POST'])
def send_forgot():
    ue= request.form["email"]
    df = pd.read_excel (filename)
    found=False
    for r in range(len(df)):
        user=str(df.iloc[r]['Username'])
        passw=str(df.iloc[r]['Password'])
        mail=str(df.iloc[r]['Email ID'])
        if mail==ue:
            found = True
            break     
    if found == True:
        msg = MIMEMultipart()
        email_html = open('email1.html')
        email_body = email_html.read()
        port = 465  # For SSL
        smtp_server = "smtp.gmail.com"
        sender_email = ""  # Enter your library mail ID
        receiver_email = mail 
        password1 = '' # Enter your library mail ID password
        if(sender_email=="" and password1=""):
            message="No user mail ID verification is done because we didn't have access to server mail ID."
            flash("Sorry we couldn't send you mail with your login credentials,because we didn't have access to server mail ID.")
        else:
            msg['Subject'] = 'Your Login credentials'
            msg['From'] = 'librarymailserver@gmail.com'
            text = MIMEText("""

            As per your request this mail was sent.

            Password can also be changed in profile.

            Login credentials which was given you on registeration.

            Username :"""+ user +"""
            Mail ID  :"""+ mail +"""
            Password :"""+ passw +"""

            This message is sent by server. Please don't reply""")
            msg.attach(MIMEText(email_body, 'html'))
            msg.attach(text)

            context = ssl.create_default_context()
            try:
                server=smtplib.SMTP_SSL(smtp_server, port, context=context) 
                server.login(sender_email, password1)
                server.sendmail(sender_email, receiver_email,msg.as_string())

            except Exception:
                print("Please connect to internet to proceed your request,No Internet Connection  / Please check that you have entered library mail ID in the server correctly ")
                flash("Please connect to internet to proceed your request,No Internet Connection / Please Check the server window for more information")
                return render_template('forgot.html')
        
            message="Please check your registered mail for login credentials"
            flash("Please check your registered mail for login credentials,message has been sent successfully.")
    else:
        message="Invalid Email ID"
        flash("Invalid Email ID / No user found in requested Email ID.Please register in register page.")
    print(message)  
    return render_template('forgot.html')


@app.route('/logout')
def logout_page():
    global un
    global e
    global password
    global d
    global yrs
    global found
    global erow
    global bookl
    global rl
    global rrow
    global rbd
    global foundr
    global rbookl
    un=''
    e=''
    d=''    
    yrs=''
    ro=''
    password='' 
    erow=[]
    found=False
    bookl=[]
    rl=[]
    rrow=[]
    rbd=[]
    rbookl=[]
    foundr=False 
        
    
    return render_template('logout.html')

@app.route('/rthome')
def return_to_homepage():
    global un
    global e
    global d
    global yrs
    if(un==''):
        return render_template('login.html')
    
    return render_template('home.html',uname=un,e=e)

@app.route('/home',methods=['POST'])
def login_process_page():
    global un
    global e
    global d
    global yrs
    global password
    global ro
    
    ue= request.form["email"]
    p= request.form["password"]
    now = datetime.datetime.now()
    login = False
    
    e=ue
    password=p
    df = pd.read_excel (filename)
    for r in range(len(df)):
        user=str(df.iloc[r]['Username'])
        passw=str(df.iloc[r]['Password'])
        mail=str(df.iloc[r]['Email ID'])
        if mail==ue and passw==p:
            un=user
            login = True
            df.at[r,'Recent login']=now
            print(df.iloc[r]['Recent login'])
            df.to_excel(filename,index=False)
            print("Existing user ")
            print("Username:",user)
            break     
    if login == True:
        message="Sucessfully logged in"
        d=df.iloc[r]['Department']
        yrs=str(df.iloc[r]['Year&Section'])
    else:
        message="User not found"
        print(message)
        flash("Invalid Username/Email ID/Password, Please check your registered mail for login credentials")
        return render_template('login.html')
    ro=r
    return render_template('home.html',uname=un,e=e)

@app.route('/rhome',methods=['POST'])
def register_homepage():
    global un
    global e
    global d
    global yrs
    global password
    uname= str(request.form["uname"])
    ue= request.form["email"]
    p= request.form["password"]
    dep=request.form["dep"]
    yr_s=request.form["YEAR&SEC"]
    now = datetime.datetime.now()
    login = True
    df = pd.read_excel (filename)
    for r in range(len(df)):
        user=str(df.iloc[r]['Username'])
        passw=str(df.iloc[r]['Password'])
        mail=str(df.iloc[r]['Email ID'])
        if user==uname and mail==ue and passw==p:
            login = False
            df.to_excel(filename,index=False)
            print("Existing user ")
            break
        elif mail==ue:
            flash('This email ID has been already registered by'+user)
            return render_template('register.html')
        elif user==uname and mail==ue :
            flash('This username and email ID has been already registered,If you have already registered Please login' )
            return render_template('login.html')
    if login==False:
        flash('Already this Username and Email ID is registered so Please Login')
        return render_template('login.html')   

    
    un=uname
    e=ue
    d=dep
    yrs=yr_s
    password=p
    refused=''
    msg = MIMEMultipart()
    email_html = open('email1.html')
    email_body = email_html.read()
    port = 465  # For SSL
    smtp_server = "smtp.gmail.com"
    sender_email = ""  # Enter your library mail ID
    receiver_email = e # Enter receiver address
    password1 = ''  # Enter your library mail ID password
    if(sender_email=="" and password1=""):
        print("No user mail ID verification is done because we didn't have access to server mail ID. \n")
    else:
        msg['Subject'] = 'Your Account is created'
        msg['From'] = '' # Enter your library mail ID
        text = MIMEText("""

        Thank you for registering.


        Login credentials which was given you on registeration.

        Username         :"""+ un +"""
        Mail ID          :"""+ receiver_email +"""
        Password         :"""+ p +"""
        Department       :"""+ dep +"""
        Year & section   :"""+ yr_s +"""      
        
        (For Staff Year &section will be empty kindly ignore)
    
        If this is not you how registered,Please ignore/delete this mail.

        This message is sent by server. Please don't reply""")
        msg.attach(MIMEText(email_body, 'html'))
        msg.attach(text)

        context = ssl.create_default_context()
        try:
            server=smtplib.SMTP_SSL(smtp_server, port, context=context) 
            server.login(sender_email, password1)
            server.sendmail(sender_email, receiver_email,msg.as_string())

        except smtplib.SMTPRecipientsRefused as e:
            print('got SMTPRecipientsRefused')
            refused = receiver_email
        except Exception:
            print("Please connect to internet to proceed your request,No Internet Connection or Please check that you have entered library mail ID in the server correctly ")
            flash("We couldn't process your request since there is no Internet Connection right now,Please connect to Internet and try again later,Thank you / Please Check the server window for more information")
            return render_template('register.html')   

        print(refused)
        if(refused!=''):
            flash("Invalid Mail ID,Please enter valid mail ID only")
            return render_template('register.html')
    print("New user ID created")
    wb = load_workbook(filename)
    ws=wb.worksheets[0]
    colummn=ws['A']
    ro=len(colummn)-1
    print(ro)
    t=tuple([uname]+[p]+[ue]+['']+[now]+[dep]+[yr_s])
    ws.append(t)
    wb.save(filename=filename)
    if login == True:
        message="Sucessfully logged in"
    else:
        message="User not found Please try again later"
        flash(message)
        print(message)
        return render_template('login.html')
   
    return render_template('home.html',uname=un,e=e)

@app.route('/profile',methods=['POST','GET'])
def profile_page():
    global un
    global e
    global d
    global yrs
    if(un==''):
        return render_template('login.html')
    yrs_str=str(yrs)
    if yrs_str=="nan":
       yrs_str="null"
    print("yrs:",yrs_str)
    return render_template('profile.html',uname=un,e=e,d=d,yrs=yrs_str)

@app.route('/profile/data1',methods=['POST'])
def profile_name_changePage():
    global un
    global ro
    yrs_str=str(yrs)
    if yrs_str=="nan":
       yrs_str="null"
    #print("yrs:",yrs_str) 
    if(un==''):
        return render_template('login.html')
    uname= request.form["uname"]
    if(str(uname)==""):
        flash('Your Username has cannot be empty and your request is rejected.')
        return render_template('profile.html',uname=un,e=e,d=d,yrs=yrs_str)

    df = pd.read_excel (filename)
    df.loc[ro,'Username']=uname
    un=uname
    df.to_excel(filename,index=False)
    flash('Your Username has been changed based on your request') 
      
    return render_template('profile.html',uname=un,e=e,d=d,yrs=yrs_str)

@app.route('/profile/data2',methods=['POST'])
def profile_dyrs_ChangePage():
    global d
    global yrs
    global ro
    if(un==''):
        return render_template('login.html')
    yrs= request.form["YEAR&SEC"]
    d= request.form["dep"]
    df = pd.read_excel (filename)
    if len(yrs)>0:
        df.loc[ro,'Year&Section']=yrs
    df.loc[ro,'Department']=d
    df.to_excel(filename,index=False)
    yrs_str=str(yrs)
    if yrs_str=="nan" or yrs_str=="":
        yrs_str="null"
        #if len(yrs)==0:
        #yrs="null"
        flash('Your Department has been changed based on your request') 
    else:
        flash('Your Department and Year and Section has been changed based on your request') 
    print("yrs:",yrs_str) 
    return render_template('profile.html',uname=un,e=e,d=d,yrs=yrs_str)

@app.route('/profile/pdata',methods=['POST'])
def profile_password_changePage():
    global password
    global ro
    global yrs
    if(un==''):
        return render_template('login.html')
    yrs_str=str(yrs)
    if yrs_str=="nan":
       yrs_str="null"
    npassword= request.form["password"] #new password
    cpassword= request.form["cupassword"] #current password
    if(len(str(npassword))==0):
        flash("You cannot change to empty password.your request is rejected.")
        return render_template('profile.html',uname=un,e=e,d=d,yrs=yrs_str)
    if(cpassword==password):
        df = pd.read_excel (filename)
        df.loc[ro,'Password']=npassword
        password=npassword
        df.to_excel(filename,index=False) 
        flash('Your Password is changed based on your request') 
    else:
        flash('Your current Password does not match with password,Please enter correctly to change')
    
    return render_template('profile.html',uname=un,e=e,d=d,yrs=yrs_str)

@app.route('/Entry')
def entry_page():
    global un
    global e
    global d
    global yrs
    if(un==''):
        return render_template('login.html')
    return render_template('Entry.html',uname=un,e=e,d=d,yrs=yrs)

@app.route('/Entry/data',methods=['POST'])
def entry_data_page():
    global un
    global e
    global d
    global yrs
    if(un==''):
        return render_template('login.html')
    Date= request.form["DATE"]
    
    books= request.form["books"]
    print(books)
    
    now = datetime.datetime.now()
    wb = load_workbook(filename1)
    ws=wb.worksheets[0]
    days_after = (date.today()+timedelta(days=30)).isoformat()  
    t=tuple([now]+[Date]+[un]+[e]+[yrs]+[d]+[books]+[days_after]+["0"]+["0"])
    ws.append(t)
    wb.save(filename=filename1)
    df = pd.read_excel(filename1)
    
    df=df.drop_duplicates(subset=["NAME", "COLLEGE MAIL ID", "YEAR & SECTION","BOOK ID","DATE"], keep="first", inplace=False)

    df.to_excel(filename1,index=False)
    flash("Your Entry is being processed and will be added in your Book Entries")
    return render_template('Entry.html',uname=un,e=e,d=d,yrs=yrs)

 
@app.route('/book')
def book_page():
    return render_template('books.html')

@app.route('/Renewal')
def renewal_page():
    global erow
    global found
    global un
    global e
    global bookl
    global rl 
    global rbd
    if(un==''):
        return render_template('login.html')

    df = pd.read_excel (filename1)
    rl=[]
    for r in range(len(df)):
        user=str(df.iloc[r]['NAME'])
        mail=str(df.iloc[r]['COLLEGE MAIL ID'])
        if user==un and mail==e:
            found = True
            books=list(map(int,df.iloc[r]['BOOK ID'].split(',')))
            print(books)
            rbooks=list(str(df.iloc[r]['BOOK RETURNED']).split(','))
            print(rbooks)
            if "0" in rbooks:
                    rbooks.remove("0")
            for i in rbooks:
                if int(i) in books:
                    books.remove(int(i))
            
            if books not in bookl and len(books)>0:
                bookl.append(books)
                erow.append(r)
                print(df.iloc[r]['NUMBER OF RENEWAL'])
                rl.append(df.iloc[r]['NUMBER OF RENEWAL'])
            
                due_date=df.iloc[r]['DUE DATE']
                rbd.append(int(due_date[8:10]))
                rbd.append(int(due_date[5:7]))
                rbd.append(int(due_date[0:4]))
            

    df.to_excel(filename1,index=False)
    print(rbd)
    
    if found==False:
        print("Cannot Find")


    return render_template('Renewal.html',uname=un,e=e,bl=bookl,rv=rl,rbdl=rbd)

@app.route('/renewal/data',methods=['POST'])
def renewal_data_page():
    global erow
    global found 
    global bookl
    global rl
    global rbd
    if(un==''):
        return render_template('login.html')

    df = pd.read_excel (filename1)
    
    if found==True:
        t=request.form["rowvalue"]
        days=int(request.form["Days"])
        print(t,days)
        books=list(map(int,t.split(',')))
        for r in bookl:
            if(books==r):
                rv=erow[bookl.index(books)]
                if(int(df.iloc[rv]['NUMBER OF RENEWAL'])>=5):
                    flash("You cannot renew this books,because maximum renewal is completed already")
                    return render_template('renewal.html',uname=un,e=e,bl=bookl,rv=rl,rbdl=rbd)
                df.at[rv,'NUMBER OF RENEWAL']=int(df.iloc[rv]['NUMBER OF RENEWAL'])+1
                print(df.iloc[rv]['DUE DATE'])
                year, month, day = map(int,df.iloc[rv]['DUE DATE'].split('-'))
                n = datetime.date(year, month, day)
                days_after = (n+timedelta(days=days)).isoformat()
              
                print(days_after)
                df.at[rv,'DUE DATE']=days_after
                rbd[(bookl.index(books))*3]=int(days_after[8:10])
                rbd[(bookl.index(books))*3+1]=int(days_after[5:7])
                rbd[(bookl.index(books))*3+2]=int(days_after[0:4])

                print("Renewal=",df.iloc[rv]['NUMBER OF RENEWAL']) 
                rl[bookl.index(books)]=df.iloc[rv]['NUMBER OF RENEWAL']  
                flash("Your Renewal request is processed and it is successfully updated")
    df.to_excel(filename1,index=False)
        
    
    if found==False:
        print("Cannot Find")


    return render_template('renewal.html',uname=un,e=e,bl=bookl,rv=rl,rbdl=rbd)

@app.route('/Book_Return')
def book_return():
    global foundr
    global un
    global e
    global rrow
    global rbookl
    rbookl=[]
    rrow=[]
    if(un==''):
        return render_template('login.html')
    

    df = pd.read_excel (filename1)
    for r in range(len(df)):
        user=str(df.iloc[r]['NAME'])
        mail=str(df.iloc[r]['COLLEGE MAIL ID'])
        if user==un and mail==e:
            foundr = True
            books=list(map(int,df.iloc[r]['BOOK ID'].split(',')))
            print(books)
            rbooks=list(str(df.iloc[r]['BOOK RETURNED']).split(','))
            if "0" in rbooks:
                rbooks.remove("0")
            print(rbooks)
            for i in rbooks:
                if int(i) in books:
                    books.remove(int(i))
            print(books)
            if books not in rbookl and len(books)>0:
                rbookl.append(books)
                rrow.append(r)
            

    df.to_excel(filename1,index=False)
        
    
    if foundr==False:
        print("Cannot Find")

    return render_template('Book_return.html',uname=un,e=e,rbl=rbookl)

@app.route('/Book_Return/data',methods=['POST'])
def book_return_data():
    global foundr
    global un
    global e
    global rrow
    global rbookl
    if(un==''):
        return render_template('login.html')
    

    df = pd.read_excel (filename1)
    #df = df.applymap(str)
    now = datetime.datetime.now()
    if foundr==True:
        t=request.form["rowvalue"]
        day=request.form["DATE"]
        staff=request.form["staff"]
        booksr=list(map(int,request.form["books"].split(",")))

        print(t,day,booksr)
        book=list(map(int,t.split(',')))
        print(rbookl,book)
        for r in range(len(rbookl)):
            if(book==rbookl[r]):
                rv=rrow[r]                
                books=list(map(int,df.iloc[rv]['BOOK ID'].split(',')))
                rbooks=list(str(df.iloc[rv]['BOOK RETURNED']).split(','))
                present=True
                for i in booksr:
                    if i not in books:
                        present=False
                print(books,rbooks,booksr)
                if present==False:
                    flash("Your Request is cancelled because you cannot return the book which is not taken by you.Please type the book ID correctly/choose check book ID correctly.")
                    return render_template('Book_return.html',uname=un,e=e,rbl=rbookl)

                rbooks=list(set(rbooks+booksr))
                print(rbooks)
                for i in rbooks:
                    if int(i) in books:
                        books.remove(int(i))
                if "0" in rbooks:
                    rbooks.remove("0")
                print(rbooks)
           
                string_value=','.join(map(str,rbooks))
                print(string_value)
                #df['BOOK RETURNED'] = df['BOOK RETURNED'].astype(str)
               
                df.at[rv,'BOOK RETURNED']=string_value
                if len(books)==0:
                    rbookl.pop(r)
                else:
                    rbookl[r]=books
                from pandas import Timestamp
                day=pd.Timestamp(day)
                
                df.at[rv,'RETURNED DATE']=day.date()
                print(day.date())
                received_staff=str(df.iloc[rv]['BOOKS RECEIVED STAFF NAME'])
                received_staffs=[i for i in received_staff.split(",")]
                if staff not in received_staff:
                    received_staffs.append(staff)
                if "nan" in received_staffs:
                    received_staffs.remove("nan")
                received_staff= ','.join(received_staffs)
                print(received_staff)
                #df['BOOKS RECEIVED STAFF NAME'] = df['BOOKS RECEIVED STAFF NAME'].astype(str)
                df.at[rv,'BOOKS RECEIVED STAFF NAME']= received_staff 
                df.at[rv,'LAST RETURN ENTRY DATE']=now       
               

                flash("Your Book Return request is processed and it is successfully updated")
    df.to_excel(filename1,index=False)
    
        
    
    if foundr==False:
        print("Cannot Find")

    return render_template('Book_return.html',uname=un,e=e,rbl=rbookl)

@app.route('/FAQ')
def faq_page():
    global un
    if(un==''):
        return render_template('login.html')
    return render_template('FAQ.html')
@app.route('/FAQ1')
def faq1_page():
    global un
    
    return render_template('FAQ1.html')

@app.route('/about')
def about_page():
    global un
    if(un==''):
        return render_template('login.html')
    return render_template('about us.html')
   

def library_line():
    if name == 'nt': 
        _ = system('cls') 
  
    # for mac and linux(here, os.name is 'posix') 
    else: 
        _ = system('clear') 

    print("\t\t\t",end=" ")
    print( "Welcome to Library Server\n")
print("\t\t\t",end=" ")
print( "Welcome to Library Server\n")
print("Please wait we will open browser for you in 2 minutes....\n")
print(r"if not opened please use http://localhost:5000/ which is Login page")
print("\n")
library_line()
http_server = WSGIServer(("localhost", 5000),app)
webbrowser.open("http://localhost:5000/")
http_server.serve_forever()
